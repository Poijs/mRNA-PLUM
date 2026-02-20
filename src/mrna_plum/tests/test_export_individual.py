from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import duckdb
import openpyxl

from mrna_plum.reports.export_individual import export_individual_reports, sanitize_filename


@dataclass
class Cfg:
    root: str
    paths: object
    reports: object


@dataclass
class Paths:
    db_path: str


@dataclass
class Reports:
    individual_dir: str = "_out/indywidualne"
    max_workers: int = 1  # tests: deterministic, no threads
    batch_teachers: int = 50
    create_empty_individual: bool = False


def _make_db(tmp_path: Path) -> duckdb.DuckDBPyConnection:
    db_path = tmp_path / "test.duckdb"
    con = duckdb.connect(str(db_path))
    con.execute("CREATE SCHEMA mart;")
    con.execute(
        """
        CREATE TABLE mart.metrics_long (
            teacher_id      VARCHAR,
            full_name       VARCHAR,
            email           VARCHAR,
            id_bazus        VARCHAR,
            course_name     VARCHAR,
            activity_label  VARCHAR,
            count_value     BIGINT,
            pct_course      DOUBLE,
            pct_kierunek    DOUBLE,
            pct_wydzial     DOUBLE,
            pct_uczelnia    DOUBLE,
            visible_active  BOOLEAN,
            hr_wydzial      VARCHAR,
            hr_jednostka    VARCHAR
        );
        """
    )
    # Teacher A: 2 rows, one has count=0 (must be excluded)
    con.execute(
        """
        INSERT INTO mart.metrics_long VALUES
        ('10', 'Kowalski Jan', 'jan@x.pl', 'B123', 'Kurs A', 'Strona', 3,  50, 10, 5, 1, TRUE, 'WL', 'Katedra X'),
        ('10', 'Kowalski Jan', 'jan@x.pl', 'B123', 'Kurs A', 'Wiki',   0,  20, 10, 5, 1, TRUE, 'WL', 'Katedra X'),
        ('10', 'Kowalski Jan', 'jan@x.pl', 'B123', 'Kurs B', 'Adres URL', 2,  25, 10, 5, 1, TRUE, 'WL', 'Katedra X');
        """
    )
    # Teacher B: only zero rows -> should be SKIPPED and no file
    con.execute(
        """
        INSERT INTO mart.metrics_long VALUES
        ('11', 'Nowak/Anna:*?', 'a@x.pl', 'B999', 'Kurs Z', 'Strona', 0, 10, 1, 1, 1, TRUE, 'WF', 'Jedn Y');
        """
    )
    return con


def test_sanitize_filename_windows_chars():
    assert sanitize_filename('Nowak/Anna:*?') == "Nowak_Anna____"


def test_export_individual_generates_one_file(tmp_path: Path):
    con = _make_db(tmp_path)
    try:
        cfg = Cfg(
            root=str(tmp_path),
            paths=Paths(db_path=str(tmp_path / "test.duckdb")),
            reports=Reports(),
        )
        code, out_dir = export_individual_reports(con, cfg)
        assert code == 0

        out_dir = Path(out_dir)
        files = sorted(out_dir.glob("*.xlsx"))
        # only teacher_id=10 should have file
        assert len(files) == 1
        assert files[0].name.endswith("_10.xlsx")
        assert "Kowalski Jan" in files[0].name

        wb = openpyxl.load_workbook(files[0])
        assert "DANE_KURSY" in wb.sheetnames
        assert "DANE_PERS" in wb.sheetnames

        ws = wb["DANE_KURSY"]
        headers = [ws.cell(1, c).value for c in range(1, 8)]
        assert headers == ["Kurs", "Aktywność", "Liczba", "% kurs", "% kierunek", "% wydział", "% uczelnia"]

        # rows: should exclude count=0, and sorted deterministically
        rows = []
        for r in range(2, ws.max_row + 1):
            rows.append([ws.cell(r, c).value for c in range(1, 8)])
        # expect two rows
        assert len(rows) == 2
        # deterministic sort: Kurs A / Strona first, Kurs B / Adres URL second
        assert rows[0][0] == "Kurs A"
        assert rows[0][1] == "Strona"
        assert rows[1][0] == "Kurs B"
        assert rows[1][1] == "Adres URL"

        # pct in XLSX: stored as 0.xx (not 50)
        # openpyxl reads raw numeric value (formatting is separate)
        assert abs(float(rows[0][3]) - 0.50) < 1e-9

        ws2 = wb["DANE_PERS"]
        # find ID_PLUM row
        kv = {ws2.cell(r, 1).value: ws2.cell(r, 2).value for r in range(2, ws2.max_row + 1)}
        assert kv["ID_PLUM"] == "10"
        assert kv["Pełna nazwa"] == "Kowalski Jan"
        assert kv["E-mail"] == "jan@x.pl"
        assert kv["ID bazus"] == "B123"
        assert kv["Wydział"] == "WL"
        assert kv["Jednostka"] == "Katedra X"

    finally:
        con.close()


def test_export_individual_skips_teacher_with_only_zero_rows(tmp_path: Path):
    con = _make_db(tmp_path)
    try:
        cfg = Cfg(
            root=str(tmp_path),
            paths=Paths(db_path=str(tmp_path / "test.duckdb")),
            reports=Reports(),
        )
        code, out_dir = export_individual_reports(con, cfg)
        assert code == 0

        out_dir = Path(out_dir)
        # verify teacher 11 is not exported
        assert not any(p.name.endswith("_11.xlsx") for p in out_dir.glob("*.xlsx"))
    finally:
        con.close()