# src/mrna_plum/rules/activity_rules.py

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


@dataclass(frozen=True)
class KeyRule:
    activity_label: str
    tech_key: str
    operation: str          # CREATE/DELETE/UPDATE/VIEW/GRADE/...
    count_mode: str         # TAK / TAK_FLAG / NIE
    rx_match: re.Pattern
    rx_user: Optional[re.Pattern]
    rx_object_from_match_group: Optional[str]  # e.g. "object_id" or None
    priority: int


@dataclass(frozen=True)
class RuleMatch:
    activity_label: str
    tech_key: str
    operation: str
    count_mode: str
    teacher_id: Optional[int]
    object_id: Optional[int]
    priority: int
    conflict: bool


def _norm(s: Any) -> str:
    return ("" if s is None else str(s)).strip()


def _compile(pat: str) -> re.Pattern:
    # wszystkie regexy traktujemy case-insensitive
    return re.compile(pat, re.IGNORECASE)


def load_keys_rules(keys_xlsx: str, sheet_name: str = "KEYS") -> List[KeyRule]:
    """
    Oczekiwane kolumny w KEYS:
    AKTYWNOSC
    KLUCZ_TECHNICZNY
    OPERACJA
    LICZYC_DO_RAPORTU
    REGEX_DOPASOWANIA_(Opis)
    REGEX_USER_ID_(Opis)
    REGEX_OBIEKT_ID_(z dopasowania)   # opcjonalnie: nazwa grupy np. object_id
    PRIORYTET
    """
    wb = load_workbook(filename=keys_xlsx, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"KEYS: brak arkusza '{sheet_name}' w {keys_xlsx}")

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = [(_norm(x)) for x in next(rows)]
    idx = {name: i for i, name in enumerate(header) if name}

    def col(name: str) -> int:
        if name not in idx:
            raise ValueError(f"KEYS: brak kolumny '{name}' (nagłówki: {header})")
        return idx[name]

    rules: List[KeyRule] = []

    for r in rows:
        activity_label = _norm(r[col("AKTYWNOSC")])
        tech_key = _norm(r[col("KLUCZ_TECHNICZNY")])
        operation = _norm(r[col("OPERACJA")]).upper()
        count_mode = _norm(r[col("LICZYC_DO_RAPORTU")]).upper()  # TAK/TAK_FLAG/NIE
        rx_match_txt = _norm(r[col("REGEX_DOPASOWANIA_(Opis)")])
        rx_user_txt = _norm(r[col("REGEX_USER_ID_(Opis)")])
        rx_obj_group = _norm(r[idx.get("REGEX_OBIEKT_ID_(z dopasowania)", -1)]) if "REGEX_OBIEKT_ID_(z dopasowania)" in idx else ""
        prio_txt = _norm(r[idx["PRIORYTET"]]) if "PRIORYTET" in idx else "1"
        if not rx_match_txt or not tech_key:
            continue

        try:
            prio = int(prio_txt) if prio_txt else 0
        except Exception:
            prio = 0

        rx_match = _compile(rx_match_txt)
        rx_user = _compile(rx_user_txt) if rx_user_txt else None
        rx_obj_group = rx_obj_group or None

        rules.append(
            KeyRule(
                activity_label=activity_label,
                tech_key=tech_key,
                operation=operation,
                count_mode=count_mode,
                rx_match=rx_match,
                rx_user=rx_user,
                rx_object_from_match_group=rx_obj_group,
                priority=prio,
            )
        )

    # najwyższy priorytet wcześniej – przyspiesza
    rules.sort(key=lambda x: x.priority, reverse=True)
    return rules


class ActivityRuleEngine:
    def __init__(self, rules: List[KeyRule], drop_mode_nie: bool = True):
        self.rules = rules
        self.drop_mode_nie = drop_mode_nie

    def match(self, opis: str) -> Optional[RuleMatch]:
        """
        Jeśli wiele reguł pasuje:
        - wybierz najwyższy PRIORYTET
        - jeśli kilka ma ten sam max PRIORYTET -> conflict=True
        """
        opis = opis or ""
        matches: List[Tuple[KeyRule, re.Match]] = []

        for rule in self.rules:
            m = rule.rx_match.search(opis)
            if m:
                matches.append((rule, m))

        if not matches:
            return None

        # wybór max priorytetu
        max_prio = max(rule.priority for rule, _ in matches)
        best = [(rule, m) for rule, m in matches if rule.priority == max_prio]

        # bierzemy pierwszy jako “winner”, ale zaznaczamy konflikt jeśli >1
        rule, m = best[0]
        conflict = len(best) > 1

        # object_id – z grupy nazwanej, jeśli KEYS wskazuje; inaczej spróbuj 1. grupy
        object_id: Optional[int] = None
        if rule.rx_object_from_match_group:
            gd = m.groupdict()
            val = gd.get(rule.rx_object_from_match_group)
            if val:
                try:
                    object_id = int(val)
                except Exception:
                    object_id = None
        else:
            try:
                if m.groups():
                    object_id = int(m.group(1))
            except Exception:
                object_id = None

        # teacher_id – regex osobny z KEYS (na Opis)
        teacher_id: Optional[int] = None
        if rule.rx_user:
            um = rule.rx_user.search(opis)
            if um:
                gd = um.groupdict()
                if "id" in gd and gd["id"]:
                    try:
                        teacher_id = int(gd["id"])
                    except Exception:
                        teacher_id = None
                else:
                    try:
                        teacher_id = int(um.group(1))
                    except Exception:
                        teacher_id = None

        if self.drop_mode_nie and rule.count_mode == "NIE":
            return None

        return RuleMatch(
            activity_label=rule.activity_label,
            tech_key=rule.tech_key,
            operation=rule.operation,
            count_mode=rule.count_mode,
            teacher_id=teacher_id,
            object_id=object_id,
            priority=rule.priority,
            conflict=conflict,
        )