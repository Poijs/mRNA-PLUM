from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from typing import List
import pandas as pd

from openpyxl import load_workbook
from ..errors import InputDataError

REQUIRED_COLS = [
    "AKTYWNOSC",
    "KLUCZ_TECHNICZNY",
    "OPERACJA",
    "LICZYC_DO_RAPORTU",
    "REGEX_DOPASOWANIA_(Opis)",
    "REGEX_USER_ID_(Opis)",
    "REGEX_OBIEKT_ID_(z dopasowania)",
    "PRIORYTET",
]

def load_keys_sheet(workbook_path: Path, sheet_name: str) -> pd.DataFrame:
    if not workbook_path.exists():
        raise InputDataError(f"KEYS workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise InputDataError(f"KEYS sheet not found: {sheet_name} in {workbook_path}")

    ws = wb[sheet_name]
    rows = list(ws.values)
    if not rows:
        raise InputDataError("KEYS sheet is empty")

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    data_rows = rows[1:]
    df = pd.DataFrame(data_rows, columns=header)

    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        raise InputDataError(f"KEYS missing columns: {missing}")

    # normalizacja
    df = df.copy()
    df["PRIORYTET"] = pd.to_numeric(df["PRIORYTET"], errors="coerce").fillna(0).astype(int)
    df["LICZYC_DO_RAPORTU"] = df["LICZYC_DO_RAPORTU"].fillna("").astype(str).str.strip()

    # usuń puste reguły
    df = df[df["KLUCZ_TECHNICZNY"].notna() & (df["KLUCZ_TECHNICZNY"].astype(str).str.strip() != "")]
    return df.reset_index(drop=True)
